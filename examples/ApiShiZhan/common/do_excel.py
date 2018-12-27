# -*- coding: UTF-8 -*-
# @Time: 2018/12/23 16:08
# @Author: Sai_12
# @Email: 932934045@qq.com
# @File: do_excel.py
'''
这个文件封装单条case和读取Excel中的数据
'''

import openpyxl
from common.request import Request

class Case:

    def __init__(self):
        self.case_id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None

class DoExcel:

    def __init__(self, file_name):
        try:
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            print("{0}没有找到，请重新检查文件路径。".format(file_name))
            raise e

    def get_cases(self, sheet_name):
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        cases = []#定义列表，盛放用例数据
        for r in range(2, max_row+1):
            case = Case()#实例化一个case对象，用来存放测试数据
            case.case_id = sheet.cell(row=r, column=1).value
            case.url = sheet.cell(row=r, column=2).value
            case.data = sheet.cell(row=r, column=3).value
            case.title = sheet.cell(row=r, column=4).value
            case.method = sheet.cell(row=r, column=5).value
            case.expected = sheet.cell(row=r, column=6).value
            cases.append(case)
        return cases

if __name__ == '__main__':
    do_excel = DoExcel('../datas/test_data2.xlsx')
    cases = do_excel.get_cases('login')
    print(cases)
    for case in cases:
        print(case.case_id)
        print(case.url)
	print("haha")
	print("公司2018-12-27")




